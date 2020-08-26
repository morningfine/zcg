import openpyxl
import os
from common.deal_path import DealPath

class DealExcel:

    def __init__(self,filename,sheet):
        self.filename = filename
        self.sheet = sheet

    def read_excel(self):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheet]
        rows = list(sh.rows)

        #获取title
        title = []
        for i in rows[0]:
            title.append(i.value)

        #获取正文
        result = []
        for item in rows[1:]:
            cont = []
            for i in item:
                cont.append(i.value)
            result.append(dict(zip(title,cont)))

        return result

    def write_one_excel(self,row,col,text):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheet]
        sh.cell(row=row,column=col,value=text)
        wb.save(self.filename)

    def write_all_excel(self,texts):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheet]
        for item in texts:
            row=item+2
            col=8
            text = texts[item]
            sh.cell(row=row,column=col,value=text)
        wb.save(self.filename)

if __name__ == '__main__':
    de = DealExcel(os.path.join(DealPath().DATA_PATH,'cases.xlsx'),'register')
    re = de.read_excel()

    print(re)
